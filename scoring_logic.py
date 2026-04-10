"""
Waterfall lead scoring: Apollo data → website scrape → DuckDuckGo fallback.
Uses Anthropic Claude 3 Haiku with strict JSON responses.
"""

from __future__ import annotations

import json
import logging
import re
from typing import TYPE_CHECKING, Any

import anthropic

if TYPE_CHECKING:
    import pandas as pd

logger = logging.getLogger(__name__)

MODEL = "claude-haiku-4-5"
# Claude Haiku 4.5 pricing
DEFAULT_COST_INPUT_PER_MILLION_USD = 1.0
DEFAULT_COST_OUTPUT_PER_MILLION_USD = 5.0

JSON_INSTRUCTION = """You must respond with a single JSON object only, no markdown fences, no other text.
Keys:
- "status": one of "YES", "NO", "MAYBE" (uppercase)
- "reason": short English explanation (one sentence)

Rules:
- YES: company clearly matches the ICP.
- NO: clearly does not match or is excluded (e.g. agency when ICP excludes agencies).
- MAYBE: insufficient or ambiguous information."""


def _extract_json(text: str) -> dict[str, Any]:
    text = text.strip()
    if not text:
        raise ValueError("empty model response")
    # Strip accidental markdown fences
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)
    data = json.loads(text)
    if "status" not in data or "reason" not in data:
        raise ValueError("missing status or reason")
    st = str(data["status"]).upper().strip()
    if st not in ("YES", "NO", "MAYBE"):
        raise ValueError(f"invalid status: {st}")
    data["status"] = st
    data["reason"] = str(data["reason"]).strip()[:2000]
    return data


def call_claude_json(
    api_key: str,
    user_content: str,
) -> tuple[dict[str, Any], int, int]:
    """
    Returns (parsed_json, input_tokens, output_tokens).
    On failure raises — caller wraps in try/except.
    """
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model=MODEL,
        max_tokens=256,
        messages=[
            {
                "role": "user",
                "content": f"{JSON_INSTRUCTION}\n\n{user_content}",
            }
        ],
    )
    text = ""
    for block in message.content:
        t = getattr(block, "text", None)
        if t:
            text += t
        elif isinstance(block, dict) and block.get("type") == "text":
            text += block.get("text") or ""
    usage = getattr(message, "usage", None)
    in_tok = int(getattr(usage, "input_tokens", 0) or 0) if usage else 0
    out_tok = int(getattr(usage, "output_tokens", 0) or 0) if usage else 0
    return _extract_json(text), in_tok, out_tok


def _safe_call_claude(
    api_key: str, prompt: str
) -> tuple[dict[str, Any] | None, int, int, str | None]:
    """Returns (data, input_tokens, output_tokens, error)."""
    try:
        data, in_tok, out_tok = call_claude_json(api_key, prompt)
        return data, in_tok, out_tok, None
    except Exception as e:
        logger.exception("Claude API error")
        return None, 0, 0, str(e)


def step1_apollo(
    api_key: str,
    icp_description: str,
    short_description: str,
    technologies: str,
    keywords: str,
) -> tuple[str, str, str, int, int]:
    """Returns (status, reason, source, input_tokens, output_tokens)."""
    has_any = any(
        str(x).strip()
        for x in (short_description, technologies, keywords)
        if x is not None and str(x).strip() and str(x).lower() != "nan"
    )
    if not has_any:
        return "MAYBE", "No Apollo description/technologies/keywords", "Apollo_Data", 0, 0

    prompt = f"""You are a senior B2B sales researcher qualifying companies for a sales pipeline.

ICP (Ideal Customer Profile) criteria:
{icp_description}

Company data from Apollo export:
- Short Description: {short_description or "(empty)"}
- Technologies: {technologies or "(empty)"}
- Keywords: {keywords or "(empty)"}

Based only on the above, does this company match the ICP?"""

    data, in_tok, out_tok, err = _safe_call_claude(api_key, prompt)
    if err or data is None:
        return "MAYBE", f"Step1 API/parse error: {err or 'unknown'}", "Apollo_Data", in_tok, out_tok

    return data["status"], data["reason"], "Apollo_Data", in_tok, out_tok


def fetch_website_text(url: str, max_chars: int = 3000) -> tuple[str | None, str | None]:
    """Fetch plain text from a URL using requests + BeautifulSoup."""
    if not url or not str(url).strip() or str(url).lower() in ("nan", "none"):
        return None, "empty URL"
    u = str(url).strip()
    if not u.startswith(("http://", "https://")):
        u = "https://" + u
    try:
        import requests
        from bs4 import BeautifulSoup

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        }
        resp = requests.get(u, headers=headers, timeout=15, allow_redirects=True)
        if resp.status_code >= 400:
            return None, f"HTTP {resp.status_code}"
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header", "form", "iframe"]):
            tag.decompose()
        text = soup.get_text(separator=" ", strip=True)
        text = re.sub(r"\s{2,}", " ", text).strip()
        if not text:
            return None, "no text extracted"
        return text[:max_chars], None
    except Exception as e:
        logger.exception("fetch_website_text error")
        return None, str(e)


def step2_website(
    api_key: str,
    icp_description: str,
    website_url: str,
) -> tuple[str, str, str, int, int]:
    """Returns (status, reason, source, input_tokens, output_tokens)."""
    text, err = fetch_website_text(website_url)
    if err or not text:
        return "MAYBE", f"Website unavailable or empty: {err or 'unknown'}", "Website_Scraped", 0, 0

    prompt = f"""You are a senior B2B sales researcher qualifying companies for a sales pipeline.

ICP criteria:
{icp_description}

Below is text extracted from the company website (may be truncated):
---
{text}
---

Does this company match the ICP?"""

    data, in_tok, out_tok, api_err = _safe_call_claude(api_key, prompt)
    if api_err or data is None:
        return "MAYBE", f"Step2 Claude error: {api_err or 'unknown'}", "Website_Scraped", in_tok, out_tok

    return data["status"], data["reason"], "Website_Scraped", in_tok, out_tok


def step3_ddg(
    api_key: str,
    icp_description: str,
    company_name: str,
    linkedin_url: str,
) -> tuple[str, str, str, int, int]:
    """Returns (status, reason, source, input_tokens, output_tokens)."""
    query = f'{company_name or ""} {linkedin_url or ""} overview'.strip()
    snippets: list[str] = []
    try:
        try:
            from ddgs import DDGS
        except ImportError:
            from duckduckgo_search import DDGS

        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=3))
        for r in results:
            title = r.get("title") or ""
            body = r.get("body") or ""
            snippets.append(f"{title}\n{body}")
    except Exception as e:
        logger.exception("duckduckgo_search error")
        return "MANUAL_REVIEW", f"DDG search failed: {e}", "DDG_Search", 0, 0

    if not snippets:
        return "MANUAL_REVIEW", "No DDG results", "DDG_Search", 0, 0

    combined = "\n\n---\n\n".join(snippets)[:4000]

    prompt = f"""You are a senior B2B sales researcher qualifying companies for a sales pipeline.

ICP criteria:
{icp_description}

DuckDuckGo search snippets for "{query}":
---
{combined}
---

Final decision: does this company match the ICP? If still unclear, answer MAYBE."""

    data, in_tok, out_tok, api_err = _safe_call_claude(api_key, prompt)
    if api_err or data is None:
        return "MANUAL_REVIEW", f"Step3 Claude error: {api_err or 'unknown'}", "DDG_Search", in_tok, out_tok

    st = data["status"]
    if st == "MAYBE":
        return "MANUAL_REVIEW", data["reason"], "DDG_Search", in_tok, out_tok
    return st, data["reason"], "DDG_Search", in_tok, out_tok


def score_company_row(
    api_key: str,
    icp_description: str,
    row: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    """
    Full waterfall for one company row (dict-like).
    Returns (result_dict with ICP_Status, Reason, Data_Source), (input_tokens, output_tokens)).
    """
    def col(*names: str) -> str:
        for n in names:
            if n in row and row[n] is not None:
                v = row[n]
                if hasattr(v, "item"):
                    try:
                        v = v.item()
                    except Exception:
                        pass
                s = str(v).strip()
                if s and s.lower() != "nan":
                    return s
        return ""

    short_desc = col("Short Description", "short_description")
    tech = col("Technologies", "technologies")
    kws = col("Keywords", "keywords")
    website = col("Website", "website", "Company Website")
    company_name = col("Company Name", "Company", "Name", "company_name")
    linkedin = col("Company Linkedin Url", "Company LinkedIn Url", "linkedin_url", "Linkedin Url")

    total_in = total_out = 0

    status, reason, source, in_tok, out_tok = step1_apollo(
        api_key, icp_description, short_desc, tech, kws
    )
    total_in += in_tok
    total_out += out_tok

    if status in ("YES", "NO"):
        return {"ICP_Status": status, "Reason": reason, "Data_Source": source}, (total_in, total_out)

    # MAYBE → step 2
    status, reason, source, in_tok, out_tok = step2_website(api_key, icp_description, website)
    total_in += in_tok
    total_out += out_tok

    if status in ("YES", "NO"):
        return {"ICP_Status": status, "Reason": reason, "Data_Source": source}, (total_in, total_out)

    # Still MAYBE or website failed → step 3
    status, reason, source, in_tok, out_tok = step3_ddg(api_key, icp_description, company_name, linkedin)
    total_in += in_tok
    total_out += out_tok

    return {"ICP_Status": status, "Reason": reason, "Data_Source": source}, (total_in, total_out)


def estimate_cost_usd(
    input_tokens: int,
    output_tokens: int,
    cost_input_per_million: float = DEFAULT_COST_INPUT_PER_MILLION_USD,
    cost_output_per_million: float = DEFAULT_COST_OUTPUT_PER_MILLION_USD,
) -> float:
    return round(
        input_tokens * cost_input_per_million / 1_000_000
        + output_tokens * cost_output_per_million / 1_000_000,
        6,
    )


def export_colored_xlsx(df: "pd.DataFrame") -> bytes:
    """Write DataFrame to xlsx and color `ICP_Status` cells (YES/NO/else)."""
    import io

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if "ICP_Status" not in headers:
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    col_idx = headers.index("ICP_Status") + 1
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = cell.value
        if val == "YES":
            cell.fill = green
        elif val == "NO":
            cell.fill = red
        else:
            cell.fill = yellow

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
